from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'demo_excel_app.xlsx'

# Create a new Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo Sheet"
    ws.append(["ID", "Name", "Age", "Occupation"])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    # Load Excel data
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    data = list(ws.values)
    return render_template('index.html', data=data)

@app.route('/add', methods=['POST'])
def add_data():
    # Get form data and write to Excel
    id = request.form['id']
    name = request.form['name']
    age = request.form['age']
    occupation = request.form['occupation']

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([id, name, age, occupation])
    wb.save(EXCEL_FILE)

    return redirect(url_for('index'))

@app.route('/download')
def download_excel():
    return send_file(EXCEL_FILE, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
